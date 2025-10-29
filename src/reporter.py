"""
Reporter - 분석 결과를 Excel 파일로 내보내기
"""

import os
from datetime import datetime
from typing import Dict, List, Any
import json

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[!] openpyxl not installed. Install it with: pip install openpyxl")
    Workbook = None


class ExcelReporter:
    """분석 결과를 Excel 파일로 생성하는 클래스"""

    def __init__(self, output_dir: str = None):
        """
        Args:
            output_dir: 엑셀 파일 저장 디렉토리
        """
        self.output_dir = output_dir or "D:/project/data-tools/outputs"
        os.makedirs(self.output_dir, exist_ok=True)

        if Workbook is None:
            print("[!] Cannot create Excel reports without openpyxl")
            self.wb = None
        else:
            self.wb = Workbook()

    def _get_styles(self):
        """스타일 정의"""
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        subheader_font = Font(name='Arial', size=10, bold=True)
        subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        subheader_alignment = Alignment(horizontal='center', vertical='center')

        data_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        return {
            'header_font': header_font,
            'header_fill': header_fill,
            'header_alignment': header_alignment,
            'subheader_font': subheader_font,
            'subheader_fill': subheader_fill,
            'subheader_alignment': subheader_alignment,
            'data_alignment': data_alignment,
            'border': border,
        }

    def _apply_cell_style(self, cell, font=None, fill=None, alignment=None, border=None):
        """셀에 스타일 적용"""
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border

    def create_summary_sheet(self, results: Dict[str, Any], detector_summary: Dict[str, Any]):
        """요약 시트 생성"""
        if self.wb is None:
            return

        ws = self.wb.active
        ws.title = 'Summary'

        styles = self._get_styles()

        row = 1

        # 제목
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = 'YOLO Detection Analysis Report'
        cell.font = Font(name='Arial', size=14, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 2

        # 생성 날짜
        ws[f'A{row}'] = 'Report Generated:'
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row += 2

        # 전체 메트릭 섹션
        ws[f'A{row}'] = 'Overall Metrics'
        self._apply_cell_style(ws[f'A{row}'], font=styles['subheader_font'], fill=styles['subheader_fill'])
        row += 1

        overall = results['metrics']['overall']
        metrics_data = [
            ['Metric', 'Value'],
            ['Precision', f"{overall['precision']:.4f}"],
            ['Recall', f"{overall['recall']:.4f}"],
            ['F1 Score', f"{overall['f1_score']:.4f}"],
            ['True Positives (TP)', str(overall['tp'])],
            ['False Positives (FP)', str(overall['fp'])],
            ['False Negatives (FN)', str(overall['fn'])],
        ]

        for idx, data_row in enumerate(metrics_data):
            ws[f'A{row}'] = data_row[0]
            ws[f'B{row}'] = data_row[1]

            if idx == 0:
                self._apply_cell_style(ws[f'A{row}'], font=styles['header_font'], fill=styles['header_fill'],
                                      alignment=styles['header_alignment'])
                self._apply_cell_style(ws[f'B{row}'], font=styles['header_font'], fill=styles['header_fill'],
                                      alignment=styles['header_alignment'])
            row += 1

        row += 1

        # Detection Rate
        ws[f'A{row}'] = 'Detection Rate (%)'
        ws[f'B{row}'] = f"{results['detection_rate']:.2f}"
        row += 2

        # 클래스별 메트릭
        if results['metrics']['per_class']:
            ws[f'A{row}'] = 'Per-Class Metrics'
            self._apply_cell_style(ws[f'A{row}'], font=styles['subheader_font'], fill=styles['subheader_fill'])
            row += 1

            class_metrics_data = [['Class', 'Precision', 'Recall', 'F1 Score', 'TP', 'FP', 'FN']]

            for class_name in sorted(results['metrics']['per_class'].keys()):
                cm = results['metrics']['per_class'][class_name]
                class_metrics_data.append([
                    class_name,
                    f"{cm['precision']:.4f}",
                    f"{cm['recall']:.4f}",
                    f"{cm['f1_score']:.4f}",
                    str(cm['tp']),
                    str(cm['fp']),
                    str(cm['fn']),
                ])

            for idx, data_row in enumerate(class_metrics_data):
                for col_idx, value in enumerate(data_row):
                    cell = ws.cell(row=row, column=col_idx+1, value=value)
                    if idx == 0:
                        self._apply_cell_style(cell, font=styles['header_font'], fill=styles['header_fill'],
                                              alignment=styles['header_alignment'])
                    else:
                        self._apply_cell_style(cell, alignment=styles['data_alignment'], border=styles['border'])
                row += 1

            row += 1

        # AP Scores
        if results['ap_scores']:
            ws[f'A{row}'] = 'Average Precision (AP)'
            self._apply_cell_style(ws[f'A{row}'], font=styles['subheader_font'], fill=styles['subheader_fill'])
            row += 1

            ap_data = [['Class', 'AP Score']]
            for class_name in sorted(results['ap_scores'].keys()):
                ap_data.append([class_name, f"{results['ap_scores'][class_name]:.4f}"])

            for idx, data_row in enumerate(ap_data):
                ws[f'A{row}'] = data_row[0]
                ws[f'B{row}'] = data_row[1]

                if idx == 0:
                    self._apply_cell_style(ws[f'A{row}'], font=styles['header_font'], fill=styles['header_fill'],
                                          alignment=styles['header_alignment'])
                    self._apply_cell_style(ws[f'B{row}'], font=styles['header_font'], fill=styles['header_fill'],
                                          alignment=styles['header_alignment'])
                row += 1

        # 열 너비 자동 조정
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def create_detection_sheet(self, detection_results: List[Dict[str, Any]]):
        """검출 결과 시트 생성"""
        if self.wb is None:
            return

        ws = self.wb.create_sheet('Detection Results')

        styles = self._get_styles()

        # 헤더
        headers = ['Image Name', 'Image Size', 'Total Detections', 'Classes', 'Confidence (Min)', 'Confidence (Max)', 'Confidence (Avg)']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_cell_style(cell, font=styles['header_font'], fill=styles['header_fill'],
                                  alignment=styles['header_alignment'])

        # 데이터
        for row_idx, detection in enumerate(detection_results, 2):
            image_name = detection.get('image_name', 'Unknown')
            image_size = detection.get('image_size', {})
            size_str = f"{image_size.get('width', 0)}x{image_size.get('height', 0)}"
            total_det = detection.get('total_detections', 0)
            classes = detection.get('unique_classes', 0)

            # 신뢰도 통계
            confidences = [d['confidence'] for d in detection.get('detections', [])]
            conf_min = min(confidences) if confidences else 0
            conf_max = max(confidences) if confidences else 0
            conf_avg = sum(confidences) / len(confidences) if confidences else 0

            row_data = [
                image_name,
                size_str,
                str(total_det),
                str(classes),
                f"{conf_min:.4f}",
                f"{conf_max:.4f}",
                f"{conf_avg:.4f}",
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                self._apply_cell_style(cell, alignment=styles['data_alignment'], border=styles['border'])

        # 열 너비
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        for col in ['C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 18

    def create_class_distribution_sheet(self, class_distribution: Dict[str, int]):
        """클래스 분포 시트 생성"""
        if self.wb is None:
            return

        ws = self.wb.create_sheet('Class Distribution')

        styles = self._get_styles()

        # 헤더
        headers = ['Class', 'Count', 'Percentage']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_cell_style(cell, font=styles['header_font'], fill=styles['header_fill'],
                                  alignment=styles['header_alignment'])

        # 데이터
        total_count = sum(class_distribution.values())
        for row_idx, (class_name, count) in enumerate(sorted(class_distribution.items(), key=lambda x: x[1], reverse=True), 2):
            percentage = (count / total_count * 100) if total_count > 0 else 0

            ws.cell(row=row_idx, column=1, value=class_name)
            ws.cell(row=row_idx, column=2, value=count)
            ws.cell(row=row_idx, column=3, value=f"{percentage:.2f}%")

            for col in range(1, 4):
                cell = ws.cell(row=row_idx, column=col)
                self._apply_cell_style(cell, alignment=styles['data_alignment'], border=styles['border'])

        # 합계
        total_row = len(class_distribution) + 2
        ws.cell(row=total_row, column=1, value='Total')
        ws.cell(row=total_row, column=2, value=total_count)

        for col in range(1, 4):
            cell = ws.cell(row=total_row, column=col)
            self._apply_cell_style(cell, font=styles['subheader_font'], fill=styles['subheader_fill'],
                                  alignment=styles['data_alignment'], border=styles['border'])

        # 열 너비
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    def create_file_type_sheet(self, file_type_distribution: Dict[str, int]):
        """파일 종류 시트 생성"""
        if self.wb is None:
            return

        ws = self.wb.create_sheet('File Types')

        styles = self._get_styles()

        # 헤더
        headers = ['File Type', 'Count', 'Percentage']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_cell_style(cell, font=styles['header_font'], fill=styles['header_fill'],
                                  alignment=styles['header_alignment'])

        # 데이터
        total_count = sum(file_type_distribution.values())
        for row_idx, (file_type, count) in enumerate(sorted(file_type_distribution.items(), key=lambda x: x[1], reverse=True), 2):
            percentage = (count / total_count * 100) if total_count > 0 else 0

            ws.cell(row=row_idx, column=1, value=file_type)
            ws.cell(row=row_idx, column=2, value=count)
            ws.cell(row=row_idx, column=3, value=f"{percentage:.2f}%")

            for col in range(1, 4):
                cell = ws.cell(row=row_idx, column=col)
                self._apply_cell_style(cell, alignment=styles['data_alignment'], border=styles['border'])

        # 열 너비
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    def create_file_classification_sheet(self, file_classification_stats: Dict[str, int],
                                        file_classification_distribution: Dict[str, int] = None):
        """파일명 분류 시트 생성"""
        if self.wb is None:
            return

        ws = self.wb.create_sheet('File Classification')

        styles = self._get_styles()

        # 헤더
        if file_classification_distribution:
            headers = ['Classification Type', 'Files Count', 'Objects Count', 'Percentage']
        else:
            headers = ['Classification Type', 'Count', 'Percentage']

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_cell_style(cell, font=styles['header_font'], fill=styles['header_fill'],
                                  alignment=styles['header_alignment'])

        # 데이터
        total_count = sum(file_classification_stats.values())
        for row_idx, (classification, count) in enumerate(sorted(file_classification_stats.items(), key=lambda x: x[1], reverse=True), 2):
            percentage = (count / total_count * 100) if total_count > 0 else 0

            ws.cell(row=row_idx, column=1, value=classification)
            ws.cell(row=row_idx, column=2, value=count)

            if file_classification_distribution:
                obj_count = file_classification_distribution.get(classification, 0)
                ws.cell(row=row_idx, column=3, value=obj_count)
                ws.cell(row=row_idx, column=4, value=f"{percentage:.2f}%")

                for col in range(1, 5):
                    cell = ws.cell(row=row_idx, column=col)
                    self._apply_cell_style(cell, alignment=styles['data_alignment'], border=styles['border'])
            else:
                ws.cell(row=row_idx, column=3, value=f"{percentage:.2f}%")

                for col in range(1, 4):
                    cell = ws.cell(row=row_idx, column=col)
                    self._apply_cell_style(cell, alignment=styles['data_alignment'], border=styles['border'])

        # 합계
        total_row = len(file_classification_stats) + 2
        ws.cell(row=total_row, column=1, value='Total')
        ws.cell(row=total_row, column=2, value=total_count)

        if file_classification_distribution:
            total_objects = sum(file_classification_distribution.values())
            ws.cell(row=total_row, column=3, value=total_objects)

            for col in range(1, 5):
                cell = ws.cell(row=total_row, column=col)
                self._apply_cell_style(cell, font=styles['subheader_font'], fill=styles['subheader_fill'],
                                      alignment=styles['data_alignment'], border=styles['border'])
        else:
            for col in range(1, 4):
                cell = ws.cell(row=total_row, column=col)
                self._apply_cell_style(cell, font=styles['subheader_font'], fill=styles['subheader_fill'],
                                      alignment=styles['data_alignment'], border=styles['border'])

        # 열 너비
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        if file_classification_distribution:
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
        else:
            ws.column_dimensions['C'].width = 15

    def save_report(self, filename: str = None) -> str:
        """엑셀 파일 저장"""
        if self.wb is None:
            print("[!] Cannot save report without openpyxl")
            return None

        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'YOLO_Analysis_Report_{timestamp}.xlsx'

        output_path = os.path.join(self.output_dir, filename)

        try:
            self.wb.save(output_path)
            print(f"[+] Excel report saved to {output_path}")
            return output_path
        except Exception as e:
            print(f"[!] Error saving Excel report: {e}")
            return None

    def generate_full_report(self, results: Dict[str, Any], detection_results: List[Dict[str, Any]] = None,
                           detector_summary: Dict[str, Any] = None, filename: str = None) -> str:
        """전체 리포트 생성"""
        print("[*] Generating Excel report...")

        # 요약 시트
        self.create_summary_sheet(results, detector_summary)

        # 검출 결과 시트
        if detection_results:
            self.create_detection_sheet(detection_results)

        # 클래스 분포 시트
        if detector_summary and 'class_distribution' in detector_summary:
            self.create_class_distribution_sheet(detector_summary['class_distribution'])

        # 파일 종류 시트
        if detector_summary and 'file_type_distribution' in detector_summary:
            self.create_file_type_sheet(detector_summary['file_type_distribution'])

        # 파일명 분류 시트
        if detector_summary and 'file_classification_stats' in detector_summary:
            self.create_file_classification_sheet(
                detector_summary['file_classification_stats'],
                detector_summary.get('file_classification_distribution', {})
            )

        # 저장
        return self.save_report(filename)


def main():
    """테스트 코드"""
    reporter = ExcelReporter()

    # 샘플 데이터
    sample_results = {
        'metrics': {
            'overall': {
                'precision': 0.85,
                'recall': 0.92,
                'f1_score': 0.88,
                'tp': 46,
                'fp': 8,
                'fn': 4,
            },
            'per_class': {
                'dog': {'precision': 0.90, 'recall': 0.95, 'f1_score': 0.92, 'tp': 19, 'fp': 2, 'fn': 1},
                'cat': {'precision': 0.80, 'recall': 0.90, 'f1_score': 0.85, 'tp': 18, 'fp': 4, 'fn': 2},
                'bird': {'precision': 0.85, 'recall': 0.85, 'f1_score': 0.85, 'tp': 9, 'fp': 2, 'fn': 1},
            },
        },
        'ap_scores': {
            'dog': 0.92,
            'cat': 0.85,
            'bird': 0.87,
        },
        'detection_rate': 92.5,
    }

    sample_detections = [
        {
            'image_name': 'image1.jpg',
            'image_size': {'width': 1920, 'height': 1080},
            'total_detections': 5,
            'unique_classes': 2,
            'detections': [
                {'class_name': 'dog', 'confidence': 0.95},
                {'class_name': 'dog', 'confidence': 0.87},
                {'class_name': 'cat', 'confidence': 0.92},
            ],
        },
    ]

    sample_summary = {
        'class_distribution': {'dog': 25, 'cat': 15, 'bird': 10},
        'file_type_distribution': {'jpg': 30, 'png': 15, 'bmp': 5},
    }

    # 리포트 생성
    report_path = reporter.generate_full_report(sample_results, sample_detections, sample_summary)
    print(f"[+] Report generated: {report_path}")


if __name__ == "__main__":
    main()
